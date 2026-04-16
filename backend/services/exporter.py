# backend/services/exporter.py
# HuntAI - AI Job Hunter Agent
# Excel export with professional styling (openpyxl)

import os
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelJobExporter:
    """Service to create premium Excel reports from job match data."""
    
    # ─── Color Palette ──────────────────────────────────────────────────────────
    HEADER_BG = "0D1117" # Dark Background from UI
    HEADER_TEXT = "FFFFFF" # White Text
    SCORE_HIGH_FILL = "D1FAE5" # Green
    SCORE_HIGH_TEXT = "065F46"
    SCORE_MID_FILL = "FEF3C7" # Amber
    SCORE_MID_TEXT = "92400E"
    SCORE_LOW_FILL = "FEE2E2" # Red
    SCORE_LOW_TEXT = "991B1B"
    ALT_ROW_FILL = "F9FAFB" # Light Gray

    def __init__(self, output_dir: str = "exports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, size=11, color=self.HEADER_TEXT)
        cell.fill = PatternFill(start_color=self.HEADER_BG, end_color=self.HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_score_style(self, cell, score: int):
        if score >= 80:
            cell.fill = PatternFill(start_color=self.SCORE_HIGH_FILL, end_color=self.SCORE_HIGH_FILL, fill_type="solid")
        elif score >= 60:
            cell.fill = PatternFill(start_color=self.SCORE_MID_FILL, end_color=self.SCORE_MID_FILL, fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=self.SCORE_LOW_FILL, end_color=self.SCORE_LOW_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    def create_report(self, run_data: Dict[str, Any], jobs: List[Dict[str, Any]]) -> str:
        """
        Generates a styled .xlsx report. 
        Returns the absolute path to the generated file.
        """
        wb = Workbook()
        
        # ─── Sheet 1: Job Results ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Job Results"
        
        headers = [
            "#", "Company", "Job Title", "Location", "Platform",
            "Match Score", "Missing Skills", "Posted At", "Apply Link",
            "AI Suggestion", "Cover Letter", "Scraped At"
        ]
        
        ws1.append(headers)
        for cell in ws1[1]:
            self._apply_header_style(cell)
            
        # Data Rows
        for i, job in enumerate(jobs, start=1):
            row_data = [
                i,
                job.get("company", ""),
                job.get("title", ""),
                job.get("location", ""),
                job.get("platform", ""),
                job.get("match_score", 0),
                ", ".join(job.get("missing_skills", [])),
                str(job.get("posted_at", "")),
                job.get("job_url", ""),
                job.get("suggestion", ""),
                job.get("cover_letter", ""),
                str(job.get("scraped_at", ""))
            ]
            ws1.append(row_data)
            
            # Apply individual styles
            row_idx = ws1.max_row
            # Score column is index 6 (F)
            self._apply_score_style(ws1.cell(row=row_idx, column=6), job.get("match_score", 0))
            # Hyperlink apply link (I)
            url_cell = ws1.cell(row=row_idx, column=9)
            url_cell.hyperlink = job.get("job_url", "")
            url_cell.font = Font(underline="single", color="0563C1")
            
            # Alt Row shading
            if i % 2 == 0:
                for j in range(1, len(headers) + 1):
                    cell = ws1.cell(row=row_idx, column=j)
                    if cell.fill.bgColor.rgb == "00000000": # No fill yet
                        cell.fill = PatternFill(start_color=self.ALT_ROW_FILL, end_color=self.ALT_ROW_FILL, fill_type="solid")

        # Column Width Adjustments
        ws1.column_dimensions['B'].width = 25 # Company
        ws1.column_dimensions['C'].width = 35 # Title
        ws1.column_dimensions['D'].width = 20 # Location
        ws1.column_dimensions['E'].width = 12 # Platform
        ws1.column_dimensions['G'].width = 30 # Missing Skills
        ws1.column_dimensions['J'].width = 40 # Suggestions
        ws1.column_dimensions['K'].width = 50 # Cover Letter
        
        ws1.freeze_panes = "A2"
        ws1.auto_filter.ref = ws1.dimensions

        # ─── Sheet 2: Summary Stats ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary Stats")
        ws2.append(["HuntAI Execution Summary"])
        ws2['A1'].font = Font(bold=True, size=14)
        
        ws2.append(["Run ID", run_data.get("id")])
        ws2.append(["Search Query", run_data.get("query")])
        ws2.append(["Location", run_data.get("location")])
        ws2.append(["Total Jobs Evaluated", len(jobs)])
        ws2.append(["Average Match Score", run_data.get("avg_match_score", 0)])
        ws2.append(["Timestamp", str(run_data.get("started_at", ""))])
        
        # ─── File Saving ────────────────────────────────────────────────────────
        filename = f"HuntAI_Results_{run_data.get('id', 'export')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return os.path.abspath(filepath)

from datetime import datetime # Late import locally for naming
